import pandas as pd
import shutil
import openpyxl

def update_excel_template(data):
    template_path = "Copy of Pranay HOME E-Bill Analysis.xlsx"
    output_path = "Generated_Solar_Report.xlsx"
    
    # Creating a copy of the template so it doesn't overwrite the original
    shutil.copy(template_path, output_path)
    
    # Load the workbook using openpyxl to keep formulas intact
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active 
    
    ws.cell(row=1, column=4).value = data.get('Consumer Name', 'N/A')
    ws.cell(row=1, column=2).value = data.get('Consumer Number', 'N/A')
    ws.cell(row=3, column=4).value = f"{data.get('Sanctioned Load', '0')} KW"
    
    wb.save(output_path)
    return output_path